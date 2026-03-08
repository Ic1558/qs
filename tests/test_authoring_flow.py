from __future__ import annotations

import unittest
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

from universal_qs_engine.api import (
    project_calc_rebuild,
    project_component_candidates_add,
    project_component_candidates_confirm,
    project_component_candidates_reject,
    project_components_add,
    project_create,
    project_export_internal,
    project_export_owner,
    project_import_drawing,
    project_members_add,
    project_members_add_typed,
    project_rates_add,
    project_review_ack,
    project_review_get,
    project_review_override,
    project_acceptance_get,
    project_acceptance_override,
)
from universal_qs_engine.geometry_engine import GeometryFallback, compute_beam_slab_intersection, compute_member_gross_volume, compute_member_net_volume
from universal_qs_engine.project_store import load_project
from unittest.mock import patch


class AuthoringFlowTests(unittest.TestCase):
    def _create_project(self, name: str = "QS Authoring Test") -> str:
        status, payload = project_create(
            {
                "name": name,
                "client": "IC",
                "site": "Bangkok",
                "factor_mode": "private",
                "overhead_rate": 0.12,
                "vat_enabled": False,
            }
        )
        self.assertEqual(status, 200)
        return payload["project"]["project_id"]

    def _add_member(
        self,
        project_id: str,
        member_code: str,
        source_ref: str = "ST-01/A",
        member_type: str = "structure_item",
    ) -> str:
        status, payload = project_members_add(
            project_id,
            {
                "member_code": member_code,
                "member_type": member_type,
                "level": "GF",
                "basis_status": "ADOPTED_DETAIL",
                "source_ref": source_ref,
            },
        )
        self.assertEqual(status, 200)
        return payload["member"]["member_id"]

    def _add_rate(self, project_id: str, code: str, context: str, mat: float, lab: float = 0.0) -> None:
        status, _ = project_rates_add(
            project_id,
            {
                "item_code": code,
                "description": code,
                "unit": "lot",
                "rate_context": context,
                "material_rate": mat,
                "labor_rate": lab,
            },
        )
        self.assertEqual(status, 200)

    def test_acceptance_evaluation_derives_correct_criteria(self) -> None:
        project_id = self._create_project("Acceptance Test")
        # Initially fails because members/components will be created with minimal data
        # Actually, a new project has 0 members, so it might pass if criteria are "no failures".
        # But let's add a member missing source_ref
        self._add_member(project_id, "M1", source_ref="")
        
        status, payload = project_acceptance_get(project_id)
        self.assertEqual(status, 200)
        self.assertFalse(payload["evaluation"]["ok"])
        self.assertFalse(payload["evaluation"]["criteria"]["audit_link_integrity"])

    def test_acceptance_override_authorizes_export(self) -> None:
        project_id = self._create_project("Acceptance Override")
        self._add_member(project_id, "M1", source_ref="")
        
        # Fails initially
        status, payload = project_acceptance_get(project_id)
        self.assertFalse(payload["evaluation"]["ok"])
        
        # Override
        status, payload = project_acceptance_override(project_id, {"justification": "Approved for testing", "author": "IC"})
        self.assertEqual(status, 200)
        self.assertTrue(payload["evaluation"]["ok"])
        self.assertTrue(payload["evaluation"]["override"]["active"])
        
        # Verify owner export is allowed (provided no block_owner flags exist)
        # In this case, missing source_ref on member DOES create a block_owner flag.
        # So owner export should STILL be blocked.
        status, response = project_export_owner(project_id)
        self.assertEqual(status, 409)
        self.assertEqual(response["error"]["code"], "owner_export_blocked")

    def test_internal_export_includes_acceptance_sheet(self) -> None:
        project_id = self._create_project("Acceptance Sheet Export")
        status, payload = project_export_internal(project_id)
        self.assertEqual(status, 200)
        workbook = load_workbook(payload["xlsx"])
        self.assertIn("Acceptance", workbook.sheetnames)

    def test_import_dxf_creates_segments_and_flags_density_fallback(self) -> None:
        project_id = self._create_project("DXF Import")
    
        mock_extraction = {
            "entities": [
                {"handle": "H1", "layer": "ST-B-G1", "type": "LINE", "length_m": 5.0},
                {"handle": "H2", "layer": "ST-S-1F", "type": "LWPOLYLINE", "area_m2": 25.0},
            ],
            "metrics": {"kept_entities": 2, "generic_entities": 0},
            "review_queue": []
        }
    
        mock_module = unittest.mock.MagicMock()
        mock_module.extract_dxf_entities.return_value = mock_extraction
        with patch.dict("sys.modules", {"universal_qs_engine.extractor_dxf": mock_module}):
            status, payload = project_import_drawing(project_id, {
                "file_path": "test.dxf",
                "scale_factor": 0.001,
                "discipline": "structure",
                "source_label": "S-101"
            })
            self.assertEqual(status, 200)
            self.assertEqual(payload["imported_segments"], 2)
        
        # Verify segments
        takeoff = payload["takeoff"]
        self.assertEqual(len(takeoff["segments"]), 2)
        # Verify members
        members = {m["member_id"]: m for m in takeoff["members"]}
        self.assertEqual(members["ST-B-G1"]["member_type"], "beam")
        self.assertEqual(members["ST-S-1F"]["member_type"], "slab")
        
        # Verify DENSITY_FALLBACK flags
        review_flags = payload["review_flags"]
        self.assertTrue(any(f["flag_type"] == "density_fallback" for f in review_flags))

    def test_import_unknown_entity_creates_candidate(self) -> None:
        project_id = self._create_project("Unknown Entity Import")
    
        mock_extraction = {
            "entities": [
                {"handle": "H3", "layer": "ARCH-FURN", "type": "INSERT", "name": "CHAIR"},
            ],
            "metrics": {"kept_entities": 1, "generic_entities": 1},
            "review_queue": []
        }
    
        mock_module = unittest.mock.MagicMock()
        mock_module.extract_dxf_entities.return_value = mock_extraction
        with patch.dict("sys.modules", {"universal_qs_engine.extractor_dxf": mock_module}):
            status, payload = project_import_drawing(project_id, {
                "file_path": "test.dxf",
                "scale_factor": 0.001,
                "discipline": "architecture",
                "source_label": "A-101"
            })
            self.assertEqual(status, 200)
            self.assertEqual(payload["imported_candidates"], 1)
        
        # Verify candidate exists in Review flags
        review_flags = payload["review_flags"]
        self.assertTrue(any(f["flag_type"] == "candidate_pending_confirmation" for f in review_flags))

    def test_import_then_override_then_rebuild(self) -> None:
        project_id = self._create_project("Import and Resolve")
        self._add_rate(project_id, "CONC", "new", 2300, 0)
    
        mock_extraction = {
            "entities": [
                {"handle": "H1", "layer": "ST-B-G1", "type": "LINE", "length_m": 5.0},
            ],
            "metrics": {"kept_entities": 1, "generic_entities": 0},
            "review_queue": []
        }
    
        mock_module = unittest.mock.MagicMock()
        mock_module.extract_dxf_entities.return_value = mock_extraction
        with patch.dict("sys.modules", {"universal_qs_engine.extractor_dxf": mock_module}):
            status, payload = project_import_drawing(project_id, {
                "file_path": "test.dxf",
                "scale_factor": 0.001,
                "discipline": "structure",
                "source_label": "S-101"
            })
            self.assertEqual(status, 200)
        
        member_id = payload["takeoff"]["members"][0]["member_id"]
        segment_id = payload["takeoff"]["segments"][0]["segment_id"]
        
        status, payload = project_calc_rebuild(project_id)
        self.assertEqual(payload["calc_graph"]["rows"][0]["basis_status"], "DENSITY_FALLBACK")
        
        # Resolve via override
        flag = next(f for f in payload["review_flags"] if f["flag_type"] == "density_fallback")
        status, payload = project_review_override(project_id, {
            "segment_id": segment_id,
            "field": "depth",
            "value": 0.4,
            "justification": "From section",
            "flag_id": flag["flag_id"]
        })
        self.assertEqual(status, 200)
        
        # Verify MANUAL_ALLOWANCE
        row = payload["calc_graph"]["rows"][0]
        self.assertEqual(row["basis_status"], "MANUAL_ALLOWANCE")
        self.assertGreater(row["qty"], 0)

    def test_owner_export_blocks_when_source_missing(self) -> None:
        project_id = self._create_project("Owner Gate")
        member_id = self._add_member(project_id, "GB1", source_ref="")
        status, _ = project_components_add(
            project_id,
            {
                "member_id": member_id,
                "component_type": "DB20",
                "spec": "DB20",
                "qty": 10,
                "unit": "kg",
                "loss_pct": 0.0,
                "line_type": "ADD",
                "rate_context": "new",
                "basis_status": "ADOPTED_DETAIL",
                "source_ref": "",
                "formula_text": "10 x 1",
            },
        )
        self.assertEqual(status, 200)

        status, response = project_export_owner(project_id)
        self.assertEqual(status, 409)
        self.assertEqual(response["error"]["code"], "owner_export_blocked")

    def test_typed_structure_member_gates_block_owner_export(self) -> None:
        project_id = self._create_project("Typed Member Gates")
        status, payload = project_members_add_typed(
            project_id,
            "beam",
            {
                "member_code": "GB1",
                "level": "GF",
                "grid_ref": "A-1 to B-1",
                "clear_span": 0,
                "section_width": 0.2,
                "section_depth": 0.4,
                "basis_status": "ADOPTED_DETAIL",
                "source_ref": "S-X1-307/GB1",
            },
        )
        self.assertEqual(status, 200)
        self.assertEqual(payload["member"]["member_type"], "beam")

        status, payload = project_members_add_typed(
            project_id,
            "slab",
            {
                "member_code": "S1",
                "level": "GF",
                "slab_type": "S1",
                "thickness": 0.15,
                "area_blocks": [{"name": "main", "area": 0}],
                "basis_status": "ADOPTED_DETAIL",
                "source_ref": "S-X1-101/S1",
            },
        )
        self.assertEqual(status, 200)
        self.assertEqual(payload["member"]["member_type"], "slab")

        status, payload = project_members_add_typed(
            project_id,
            "pedestal",
            {
                "member_code": "PED-03",
                "level": "GF",
                "type_ref": "ตอม่อ 3",
                "H_to_top_of_beam": 1.2,
                "basis_status": "ADOPTED_DETAIL",
                "source_ref": "S-X1-103/ตอม่อ 3",
            },
        )
        self.assertEqual(status, 200)
        self.assertEqual(payload["member"]["member_type"], "pedestal")

        status, payload = project_review_get(project_id)
        self.assertEqual(status, 200)
        flag_types = {flag["flag_type"] for flag in payload["review_flags"]}
        self.assertIn("beam_span_unclosed", flag_types)
        self.assertIn("slab_area_unclosed", flag_types)
        self.assertIn("pedestal_h_unclosed", flag_types)

        status, response = project_export_owner(project_id)
        self.assertEqual(status, 409)
        self.assertEqual(response["error"]["code"], "owner_export_blocked")

    def test_abt_override_propagates_to_engine_total(self) -> None:
        project_id = self._create_project("ABT Override")
        self._add_rate(project_id, "ABT_01", "abortive", 57440, 0)
        member_id = self._add_member(project_id, "A-05")

        status, _ = project_components_add(
            project_id,
            {
                "member_id": member_id,
                "component_type": "ABT_01",
                "spec": "ABT_01",
                "qty": 1,
                "unit": "lot",
                "loss_pct": 0.0,
                "line_type": "ABT",
                "rate_context": "abortive",
                "basis_status": "NEGOTIATED_COMMERCIAL",
                "source_ref": "GF-ST-03/A-05",
                "abt_charged_override": 40000,
                "formula_text": "negotiated abortive",
            },
        )
        self.assertEqual(status, 200)

        status, _ = project_calc_rebuild(project_id)
        self.assertEqual(status, 200)
        status, payload = project_export_internal(project_id)
        self.assertEqual(status, 200)
        self.assertEqual(payload["summary"]["D"], 40000)
        self.assertEqual(payload["summary"]["final_bid"], 44800)
        self.assertTrue(Path(payload["xlsx"]).exists())

    def test_kamolpat_final_bid_reproduced_from_project_state(self) -> None:
        project_id = self._create_project("Kamolpat Project State")

        self._add_rate(project_id, "K_ADD", "new", 389512, 0)
        self._add_rate(project_id, "K_ABT", "abortive", 57440, 0)
        self._add_rate(project_id, "K_CRD", "reuse_recovery", 144388, 0)

        add_member_id = self._add_member(project_id, "VO.ADD")
        abt_member_id = self._add_member(project_id, "VO.ABT")
        crd_member_id = self._add_member(project_id, "VO.CRD")

        inputs = [
            {
                "member_id": add_member_id,
                "component_type": "K_ADD",
                "spec": "K_ADD",
                "qty": 1,
                "unit": "lot",
                "loss_pct": 0.0,
                "line_type": "ADD",
                "rate_context": "new",
                "basis_status": "VERIFIED_DETAIL",
                "source_ref": "ST-NEW/ADD",
                "formula_text": "base add total",
            },
            {
                "member_id": abt_member_id,
                "component_type": "K_ABT",
                "spec": "K_ABT",
                "qty": 1,
                "unit": "lot",
                "loss_pct": 0.0,
                "line_type": "ABT",
                "rate_context": "abortive",
                "basis_status": "NEGOTIATED_COMMERCIAL",
                "source_ref": "ST-OLD/ABT",
                "abt_charged_override": 40000,
                "formula_text": "actual 57440 charged 40000",
            },
            {
                "member_id": crd_member_id,
                "component_type": "K_CRD",
                "spec": "K_CRD",
                "qty": -1,
                "unit": "lot",
                "loss_pct": 0.0,
                "line_type": "CRD",
                "rate_context": "reuse_recovery",
                "basis_status": "VERIFIED_DETAIL",
                "source_ref": "ST-OLD/CRD",
                "formula_text": "credit net",
            },
        ]
        for payload in inputs:
            status, _ = project_components_add(project_id, payload)
            self.assertEqual(status, 200)

        status, _ = project_calc_rebuild(project_id)
        self.assertEqual(status, 200)
        status, payload = project_export_owner(project_id)
        self.assertEqual(status, 200)
        self.assertEqual(payload["summary"]["D"], 285124)
        self.assertEqual(payload["summary"]["final_bid"], 319339)
        self.assertTrue(Path(payload["xlsx"]).exists())

    def test_internal_export_contains_trace_layers(self) -> None:
        project_id = self._create_project("Trace Workbook")
        self._add_rate(project_id, "DB20", "new", 44, 12)
        member_id = self._add_member(project_id, "GB5")
        status, _ = project_components_add(
            project_id,
            {
                "member_id": member_id,
                "component_type": "DB20",
                "spec": "DB20",
                "qty": 100,
                "unit": "kg",
                "line_type": "ADD",
                "rate_context": "new",
                "basis_status": "VERIFIED_DETAIL",
                "source_ref": "ST-01/GB5",
            },
        )
        self.assertEqual(status, 200)
        status, payload = project_export_internal(project_id)
        self.assertEqual(status, 200)
        workbook_path = Path(payload["xlsx"])
        self.assertTrue(workbook_path.exists())
        with ZipFile(workbook_path) as zf:
            workbook_xml = zf.read("xl/workbook.xml").decode("utf-8")
        for sheet_name in ("Raw_Members", "Raw_Segments", "Raw_Components", "Raw_Candidates", "Calc_Rows", "Link_Sheet", "Detail_BOQ", "Summary", "Review"):
            self.assertIn(sheet_name, workbook_xml)

    def test_internal_workbook_formula_chain_uses_segment_refs(self) -> None:
        project_id = self._create_project("Formula Chain")
        self._add_rate(project_id, "CONC", "new", 2300, 0)
        status, payload = project_members_add_typed(
            project_id,
            "beam",
            {
                "member_code": "GB2",
                "level": "GF",
                "grid_ref": "1-2",
                "clear_span": 5.0,
                "section_width": 0.2,
                "section_depth": 0.4,
                "basis_status": "VERIFIED_DETAIL",
                "source_ref": "S-X1-307/GB2",
            },
        )
        self.assertEqual(status, 200)
        member_id = payload["member"]["member_id"]

        from universal_qs_engine.takeoff_workspace import add_segment

        segment = add_segment(
            project_id,
            {
                "member_id": member_id,
                "segment_name": "main_span",
                "length": 5.0,
                "width": 0.2,
                "depth": 0.4,
                "height": 0.4,
                "area": 0.0,
                "basis_status": "VERIFIED_DETAIL",
                "source_ref": "S-X1-307/GB2",
            },
        )

        status, _ = project_components_add(
            project_id,
            {
                "member_id": member_id,
                "source_segment_id": segment["segment_id"],
                "component_type": "CONC",
                "spec": "Beam concrete",
                "qty": 0,
                "unit": "m3",
                "line_type": "ADD",
                "rate_context": "new",
                "basis_status": "VERIFIED_DETAIL",
                "source_ref": "S-X1-307/GB2",
            },
        )
        self.assertEqual(status, 200)

        status, payload = project_export_internal(project_id)
        self.assertEqual(status, 200)
        workbook = load_workbook(payload["xlsx"], data_only=False)
        self.assertEqual(workbook["Raw_Segments"]["I2"].value, "=D2*E2*F2")
        self.assertEqual(workbook["Raw_Segments"]["M2"].value, "=J2*K2*L2")
        self.assertEqual(workbook["Calc_Rows"]["F2"].value, "='Raw_Segments'!M2")
        self.assertEqual(workbook["Link_Sheet"]["C2"].value, "=Detail_BOQ!N2")
        self.assertEqual(workbook["Link_Sheet"]["D2"].value, 2)
        self.assertEqual(workbook["Summary"]["B5"].value, "=B9")
        self.assertEqual(workbook["Summary"]["B7"].value, "=B5*B6")

    def test_dim_override_upgrades_basis_status(self) -> None:
        project_id = self._create_project("Override Flow")
        status, payload = project_members_add_typed(
            project_id,
            "beam",
            {
                "member_code": "GB11",
                "level": "GF",
                "grid_ref": "4-5",
                "clear_span": 5.0,
                "section_width": 0.2,
                "section_depth": 0.4,
                "basis_status": "ADOPTED_DETAIL",
                "source_ref": "S-X1-307/GB11",
            },
        )
        self.assertEqual(status, 200)
        member_id = payload["member"]["member_id"]
        from universal_qs_engine.takeoff_workspace import add_segment
        segment = add_segment(
            project_id,
            {
                "member_id": member_id,
                "segment_name": "fallback_zone",
                "length": 5.0,
                "width": 0.2,
                "depth": 0.0,
                "basis_status": "ADOPTED_DETAIL",
                "source_ref": "S-X1-307/GB11",
            },
        )
        status, _ = project_components_add(
            project_id,
            {
                "member_id": member_id,
                "source_segment_id": segment["segment_id"],
                "component_type": "CONC",
                "spec": "Beam concrete",
                "qty": 0,
                "unit": "m3",
                "line_type": "ADD",
                "rate_context": "new",
                "basis_status": "ADOPTED_DETAIL",
                "source_ref": "S-X1-307/GB11",
            },
        )
        self.assertEqual(status, 200)
        status, payload = project_calc_rebuild(project_id)
        self.assertEqual(status, 200)
        self.assertEqual(payload["calc_graph"]["rows"][0]["basis_status"], "DENSITY_FALLBACK")

        status, payload = project_review_override(
            project_id,
            {
                "segment_id": segment["segment_id"],
                "field": "depth",
                "value": 0.3,
                "justification": "Confirmed from section",
            },
        )
        self.assertEqual(status, 200)
        row = payload["calc_graph"]["rows"][0]
        self.assertEqual(row["basis_status"], "MANUAL_ALLOWANCE")
        self.assertAlmostEqual(row["qty"], 0.3, places=6)

        status, export_payload = project_export_internal(project_id)
        self.assertEqual(status, 200)
        workbook = load_workbook(export_payload["xlsx"], data_only=False)
        self.assertEqual(workbook["Raw_Segments"]["L2"].value, 0.3)
        self.assertEqual(workbook["Raw_Segments"]["N2"].value, "depth=0.3 (Confirmed from section)")

    def test_ack_note_does_not_unblock_owner_export(self) -> None:
        project_id = self._create_project("Ack Does Not Unlock")
        status, payload = project_members_add_typed(
            project_id,
            "beam",
            {
                "member_code": "GB12",
                "level": "GF",
                "grid_ref": "5-6",
                "clear_span": 0,
                "section_width": 0.2,
                "section_depth": 0.4,
                "basis_status": "ADOPTED_DETAIL",
                "source_ref": "S-X1-307/GB12",
            },
        )
        self.assertEqual(status, 200)
        status, review_payload = project_review_get(project_id)
        self.assertEqual(status, 200)
        beam_flag = next(flag for flag in review_payload["review_flags"] if flag["flag_type"] == "beam_span_unclosed")
        status, ack_payload = project_review_ack(
            project_id,
            {
                "flag_id": beam_flag["flag_id"],
                "comment": "Waiting for engineer reply",
            },
        )
        self.assertEqual(status, 200)
        acked_flag = next(flag for flag in ack_payload["review_flags"] if flag["flag_id"] == beam_flag["flag_id"])
        self.assertEqual(acked_flag["resolution_status"], "acknowledged")
        self.assertEqual(acked_flag["ack_comment"], "Waiting for engineer reply")

        status, response = project_export_owner(project_id)
        self.assertEqual(status, 409)
        self.assertEqual(response["error"]["code"], "owner_export_blocked")

    def test_candidate_confirm_with_reason_includes_in_calc(self) -> None:
        project_id = self._create_project("Candidate Confirm Reason")
        self._add_rate(project_id, "DB16", "new", 44, 0)
        member_id = self._add_member(project_id, "GB13")
        status, payload = project_component_candidates_add(
            project_id,
            {
                "member_id": member_id,
                "component_type": "DB16",
                "spec": "DB16",
                "qty": 6,
                "unit": "kg",
                "line_type": "ADD",
                "rate_context": "new",
                "basis_status": "ADOPTED_DETAIL",
                "source_ref": "AI-SUGGEST/GB13",
            },
        )
        self.assertEqual(status, 200)
        candidate_id = payload["candidate"]["candidate_id"]
        status, payload = project_component_candidates_confirm(project_id, candidate_id, {"reason": "QS confirmed from plan"})
        self.assertEqual(status, 200)
        self.assertEqual(payload["candidate"]["resolution_reason"], "QS confirmed from plan")
        self.assertEqual(payload["candidate"]["confirmation_status"], "confirmed")
        self.assertEqual(len(payload["calc_graph"]["rows"]), 1)

    def test_candidate_reject_removes_from_calc(self) -> None:
        project_id = self._create_project("Candidate Reject")
        member_id = self._add_member(project_id, "GB14")
        status, payload = project_component_candidates_add(
            project_id,
            {
                "member_id": member_id,
                "component_type": "DB12",
                "spec": "DB12",
                "qty": 5,
                "unit": "kg",
                "line_type": "ADD",
                "rate_context": "new",
                "basis_status": "ADOPTED_DETAIL",
                "source_ref": "AI-SUGGEST/GB14",
            },
        )
        self.assertEqual(status, 200)
        candidate_id = payload["candidate"]["candidate_id"]
        status, payload = project_component_candidates_reject(project_id, candidate_id, {"reason": "Wrong member mapping"})
        self.assertEqual(status, 200)
        self.assertEqual(payload["candidate"]["confirmation_status"], "rejected")
        self.assertEqual(payload["candidate"]["resolution_reason"], "Wrong member mapping")
        self.assertEqual(len(payload["calc_graph"]["rows"]), 0)
        status, review_payload = project_review_get(project_id)
        self.assertEqual(status, 200)
        self.assertFalse(any(flag["target_ref"] == candidate_id for flag in review_payload["review_flags"]))

    def test_calc_graph_volume_component_uses_segment_volume(self) -> None:
        project_id = self._create_project("Segment Volume")
        status, payload = project_members_add_typed(
            project_id,
            "beam",
            {
                "member_code": "GB3",
                "level": "GF",
                "grid_ref": "2-3",
                "clear_span": 5.0,
                "section_width": 0.2,
                "section_depth": 0.4,
                "basis_status": "VERIFIED_DETAIL",
                "source_ref": "S-X1-307/GB3",
            },
        )
        self.assertEqual(status, 200)
        member_id = payload["member"]["member_id"]

        from universal_qs_engine.takeoff_workspace import add_segment

        segment = add_segment(
            project_id,
            {
                "member_id": member_id,
                "segment_name": "main_span",
                "length": 5.0,
                "width": 0.2,
                "depth": 0.4,
                "height": 0.4,
                "basis_status": "VERIFIED_DETAIL",
                "source_ref": "S-X1-307/GB3",
            },
        )
        status, _ = project_components_add(
            project_id,
            {
                "member_id": member_id,
                "source_segment_id": segment["segment_id"],
                "component_type": "CONC",
                "spec": "Beam concrete",
                "qty": 0,
                "unit": "m3",
                "line_type": "ADD",
                "rate_context": "new",
                "basis_status": "VERIFIED_DETAIL",
                "source_ref": "S-X1-307/GB3",
            },
        )
        self.assertEqual(status, 200)

        status, payload = project_calc_rebuild(project_id)
        self.assertEqual(status, 200)
        row = payload["calc_graph"]["rows"][0]
        self.assertAlmostEqual(row["qty"], 0.4, places=6)
        self.assertEqual(row["source_segment_id"], segment["segment_id"])

    def test_calc_graph_volume_component_falls_back_on_incomplete_geometry(self) -> None:
        project_id = self._create_project("Geometry Fallback")
        status, payload = project_members_add_typed(
            project_id,
            "beam",
            {
                "member_code": "GB4",
                "level": "GF",
                "grid_ref": "3-4",
                "clear_span": 5.0,
                "section_width": 0.2,
                "section_depth": 0.4,
                "basis_status": "VERIFIED_DETAIL",
                "source_ref": "S-X1-307/GB4",
            },
        )
        self.assertEqual(status, 200)
        member_id = payload["member"]["member_id"]

        from universal_qs_engine.takeoff_workspace import add_segment

        segment = add_segment(
            project_id,
            {
                "member_id": member_id,
                "segment_name": "incomplete_span",
                "length": 5.0,
                "width": 0.2,
                "depth": 0.0,
                "height": 0.4,
                "basis_status": "ADOPTED_DETAIL",
                "source_ref": "S-X1-307/GB4",
            },
        )
        status, _ = project_components_add(
            project_id,
            {
                "member_id": member_id,
                "source_segment_id": segment["segment_id"],
                "component_type": "CONC",
                "spec": "Beam concrete",
                "qty": 0,
                "unit": "m3",
                "line_type": "ADD",
                "rate_context": "new",
                "basis_status": "VERIFIED_DETAIL",
                "source_ref": "S-X1-307/GB4",
            },
        )
        self.assertEqual(status, 200)

        status, payload = project_calc_rebuild(project_id)
        self.assertEqual(status, 200)
        row = payload["calc_graph"]["rows"][0]
        self.assertEqual(row["basis_status"], "DENSITY_FALLBACK")

    def test_geometry_engine_basic_volume_and_fallback(self) -> None:
        gross = compute_member_gross_volume({}, [{"length": 5.0, "width": 0.2, "depth": 0.4}])
        self.assertAlmostEqual(gross, 0.4, places=6)
        net = compute_member_net_volume({}, [{"length": 5.0, "width": 0.2, "depth": 0.4}], [0.1])
        self.assertAlmostEqual(net, 0.3, places=6)
        try:
            intersection = compute_beam_slab_intersection(
                {"length": 5.0, "width": 0.2, "depth": 0.4},
                {"length": 5.0, "width": 1.0, "depth": 0.15},
            )
        except GeometryFallback:
            return
        self.assertGreater(intersection, 0.0)

    def test_pedestal_owner_export_blocks_until_h_closed(self) -> None:
        project_id = self._create_project("Pedestal Gate")
        member_id = self._add_member(project_id, "PED-01", source_ref="ST-03/PED")
        # Replace member type through direct member add path with pedestal
        # Create an explicit pedestal member for gate behavior.
        status, payload = project_members_add(
            project_id,
            {
                "member_code": "PED-02",
                "member_type": "pedestal",
                "level": "GF",
                "basis_status": "ADOPTED_DETAIL",
                "source_ref": "ST-03/PED-02",
            },
        )
        self.assertEqual(status, 200)
        pedestal_member_id = payload["member"]["member_id"]
        self._add_rate(project_id, "DB16", "new", 44, 12)
        status, _ = project_components_add(
            project_id,
            {
                "member_id": pedestal_member_id,
                "component_type": "DB16",
                "spec": "DB16",
                "qty": 25,
                "unit": "kg",
                "line_type": "ADD",
                "rate_context": "new",
                "basis_status": "ADOPTED_DETAIL",
                "source_ref": "ST-03/PED-02",
            },
        )
        self.assertEqual(status, 200)
        status, response = project_export_owner(project_id)
        self.assertEqual(status, 409)
        self.assertEqual(response["error"]["code"], "owner_export_blocked")

    def test_candidate_component_excluded_until_confirmed(self) -> None:
        project_id = self._create_project("Candidate Boundary")
        self._add_rate(project_id, "DB20", "new", 44, 0)
        member_id = self._add_member(project_id, "GB9")

        status, _ = project_components_add(
            project_id,
            {
                "member_id": member_id,
                "component_type": "DB20",
                "spec": "DB20",
                "qty": 10,
                "unit": "kg",
                "line_type": "ADD",
                "rate_context": "new",
                "basis_status": "VERIFIED_DETAIL",
                "source_ref": "ST-01/GB9",
            },
        )
        self.assertEqual(status, 200)

        status, payload = project_component_candidates_add(
            project_id,
            {
                "member_id": member_id,
                "component_type": "DB20",
                "spec": "DB20",
                "qty": 5,
                "unit": "kg",
                "line_type": "ADD",
                "rate_context": "new",
                "basis_status": "ADOPTED_DETAIL",
                "source_ref": "AI-SUGGEST/GB9",
                "candidate_source": "ai_assist",
                "ai_origin": {"provider": "openai", "model": "gpt-5"},
            },
        )
        self.assertEqual(status, 200)
        candidate_id = payload["candidate"]["candidate_id"]

        status, rebuild_payload = project_calc_rebuild(project_id)
        self.assertEqual(status, 200)
        self.assertEqual(len(rebuild_payload["calc_graph"]["rows"]), 1)
        self.assertTrue(
            any(flag["flag_type"] == "candidate_pending_confirmation" for flag in rebuild_payload["review_flags"])
        )
    
        # Phase 5: must override acceptance because of pending candidate
        project_acceptance_override(project_id, {"justification": "Test bypass", "author": "test"})

        status, owner_payload = project_export_owner(project_id)
        self.assertEqual(status, 200)
        self.assertEqual(owner_payload["summary"]["D"], 440)

        status, payload = project_component_candidates_confirm(project_id, candidate_id)
        self.assertEqual(status, 200)
        self.assertEqual(payload["candidate"]["confirmation_status"], "confirmed")

        status, rebuild_payload = project_calc_rebuild(project_id)
        self.assertEqual(status, 200)
        self.assertEqual(len(rebuild_payload["calc_graph"]["rows"]), 2)

        status, owner_payload = project_export_owner(project_id)
        self.assertEqual(status, 200)
        self.assertEqual(owner_payload["summary"]["D"], 660)

    def test_dirty_flags_clear_after_rebuild(self) -> None:
        project_id = self._create_project("Dirty Flag Flow")
        self._add_rate(project_id, "RB9", "new", 25, 15)
        member_id = self._add_member(project_id, "GB10")
        status, _ = project_components_add(
            project_id,
            {
                "member_id": member_id,
                "component_type": "RB9",
                "spec": "RB9",
                "qty": 20,
                "unit": "kg",
                "line_type": "ADD",
                "rate_context": "new",
                "basis_status": "ADOPTED_DETAIL",
                "source_ref": "ST-01/GB10",
            },
        )
        self.assertEqual(status, 200)

        project = load_project(project_id)
        self.assertTrue(project["takeoff"]["members"][0]["dirty"])
        self.assertTrue(project["takeoff"]["components"][0]["dirty"])

        status, _ = project_calc_rebuild(project_id)
        self.assertEqual(status, 200)

        project = load_project(project_id)
        self.assertFalse(project["takeoff"]["members"][0]["dirty"])
        self.assertFalse(project["takeoff"]["components"][0]["dirty"])
        self.assertFalse(project["calc_graph"]["dirty_all"])


if __name__ == "__main__":
    unittest.main()
