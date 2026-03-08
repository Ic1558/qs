from __future__ import annotations

import os
import sys
import json
from pathlib import Path

# Add src to sys.path
ROOT = Path(__file__).resolve().parents[0]
sys.path.insert(0, str(ROOT / "src"))

from universal_qs_engine.api import (
    project_create,
    project_members_add_typed,
    project_aggregate,
    project_rates_add,
    project_acceptance_get,
    project_export_internal,
    project_acceptance_override,
    project_export_owner
)
from universal_qs_engine.project_store import load_project
from openpyxl import load_workbook

def run_proof():
    print("--- Phase 6 Real Project Proof ---")
    
    # 1. Create Project
    status, payload = project_create({
        "name": "Phase 6 Real Proof",
        "client": "Wei",
        "site": "Sukhumvit 49",
        "factor_mode": "private"
    })
    prj_id = payload["project"]["project_id"]
    print(f"Created project: {prj_id}")

    # 2. Add AR Members
    # Wall
    status, payload = project_members_add_typed(prj_id, "wall", {
        "member_code": "W1",
        "level": "1F",
        "wall_type": "Brick 10cm",
        "height": 3.0,
        "gross_area": 30.0,
        "basis_status": "VERIFIED_DETAIL",
        "source_ref": "A-101/W1"
    })
    wall_id = payload["member"]["member_id"]
    print(f"Added Wall: {wall_id}")

    # Opening
    status, payload = project_members_add_typed(prj_id, "opening", {
        "member_code": "D1",
        "parent_wall_id": wall_id,
        "opening_type": "Door",
        "width": 1.0,
        "height": 2.0,
        "area": 2.0,
        "count": 1,
        "basis_status": "VERIFIED_DETAIL",
        "source_ref": "A-101/D1"
    })
    print("Added Opening")

    # Finish
    status, payload = project_members_add_typed(prj_id, "finish", {
        "member_code": "F1",
        "parent_member_id": wall_id,
        "finish_type": "Paint",
        "basis_status": "ADOPTED_DETAIL",
        "source_ref": "A-101/F1"
    })
    print("Added Finish")

    # 3. Add MEP Members
    # Count
    status, payload = project_members_add_typed(prj_id, "mep_count", {
        "member_code": "EL-SK-01",
        "level": "1F",
        "item_type": "Socket",
        "count": 12,
        "basis_status": "VERIFIED_DETAIL",
        "source_ref": "E-101/SK"
    })
    print("Added MEP Count Item")

    # Run
    status, payload = project_members_add_typed(prj_id, "mep_run", {
        "member_code": "P-CW-01",
        "level": "1F",
        "service_type": "Cold Water Pipe",
        "length": 45.5, # VALID
        "basis_status": "VERIFIED_DETAIL",
        "source_ref": "P-101/CW"
    })
    print("Added MEP Run Item (valid length)")

    # 4. Add Rates
    project_rates_add(prj_id, {"item_code": "Brick 10cm", "material_rate": 450, "labor_rate": 120, "unit": "m2"})
    project_rates_add(prj_id, {"item_code": "Paint", "material_rate": 80, "labor_rate": 45, "unit": "m2"})
    project_rates_add(prj_id, {"item_code": "Socket", "material_rate": 150, "labor_rate": 65, "unit": "set"})
    project_rates_add(prj_id, {"item_code": "Cold Water Pipe", "material_rate": 220, "labor_rate": 85, "unit": "m"})
    print("Added Rates")

    # 4.5 Add a pending candidate (Should block acceptance but NOT flags)
    from universal_qs_engine.api import project_component_candidates_add
    project_component_candidates_add(prj_id, {
        "member_id": wall_id,
        "component_type": "ACCESSORY",
        "qty": 1,
        "unit": "set",
        "source_ref": "AI-DETECTION"
    })
    print("Added Pending Candidate")

    # 5. Aggregate
    print("Running aggregator...")
    status, payload = project_aggregate(prj_id)
    if status != 200:
        print(f"Aggregation failed: {payload}")
        return

    # 6. Verify Acceptance Checklist
    print("Verifying Acceptance...")
    status, payload = project_acceptance_get(prj_id)
    evaluation = payload["evaluation"]
    print(f"Overall Acceptance: {evaluation['ok']}")
    for k, v in evaluation["criteria"].items():
        print(f"  - {k}: {v}")
    
    if not evaluation["ok"]:
        print(f"  Summary: {evaluation['summary']}")
        # Debug: find what is blocking
        prj = load_project(prj_id)
        blocking = [f for f in prj.get("review_flags", []) if f.get("severity") == "block_owner" and f.get("resolution_status") != "resolved"]
        if blocking:
            print("  Blocking Flags:")
            for b in blocking:
                print(f"    - {b.get('flag_type')}: {b.get('message')} (target: {b.get('target_ref')})")

    # 7. Export Internal + Verify Acceptance Sheet
    print("Exporting internal workbook...")
    status, payload = project_export_internal(prj_id)
    xlsx_path = payload["xlsx"]
    print(f"XLSX Path: {xlsx_path}")
    
    wb = load_workbook(xlsx_path)
    if "Acceptance" in wb.sheetnames:
        print("✅ Acceptance sheet found in workbook")
        ws = wb["Acceptance"]
        print(f"  Acceptance Status in sheet: {ws['B3'].value}")
    else:
        print("❌ Acceptance sheet MISSING from workbook")

    # 8. Test Owner Export Block
    print("Attempting owner export (should block if criteria fail)...")
    status, payload = project_export_owner(prj_id)
    if status == 409:
        print(f"✅ Owner export correctly blocked: {payload['error']['code']}")
    else:
        print(f"❌ Owner export NOT blocked (status {status})")

    # 9. Apply Override
    print("Applying Acceptance Override...")
    status, payload = project_acceptance_override(prj_id, {
        "justification": "Phase 6 proof manual override",
        "author": "Codex"
    })
    print(f"Override OK: {payload['evaluation']['ok']}")

    # 10. Verify Owner Export Allowed
    print("Attempting owner export again...")
    status, payload = project_export_owner(prj_id)
    if status == 200:
        print("✅ Owner export allowed after override")
    else:
        print(f"❌ Owner export still blocked (status {status}): {payload}")

if __name__ == "__main__":
    run_proof()
