from __future__ import annotations

from pathlib import Path
from typing import Any, Dict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .project_store import load_project
from .review_engine import rebuild_review_flags


HEADER_FILL = PatternFill("solid", fgColor="E7EFE8")
SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")
THIN = Side(style="thin", color="888888")
FONT = Font(name="Calibri", size=11, color="000000")
FONT_BOLD = Font(name="Calibri", size=11, bold=True, color="000000")


def _style_header(ws, row_idx: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = FONT_BOLD
        cell.fill = HEADER_FILL
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 40)


def _write_table(ws, headers: list[str], rows: list[list[Any]]) -> None:
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=header)
    _style_header(ws, 1, len(headers))
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
            ws.cell(row=r_idx, column=c_idx).font = FONT
    ws.freeze_panes = "A2"
    _auto_width(ws)


def write_internal_workbook(project_id: str, output_path: str, summary: Dict[str, Any]) -> str:
    project = load_project(project_id)
    review_flags = rebuild_review_flags(project_id)
    seg_row_map = {seg["segment_id"]: idx for idx, seg in enumerate(project["takeoff"]["segments"], start=2)}
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    raw_members = wb.create_sheet("Raw_Members")
    _write_table(
        raw_members,
        [
            "member_id",
            "discipline",
            "member_type",
            "member_code",
            "level",
            "basis_status",
            "execution_status",
            "grid_ref",
            "clear_span",
            "section_width",
            "section_depth",
            "slab_type",
            "thickness",
            "type_ref",
            "H_to_top_of_beam",
            "source_ref",
            "notes",
        ],
        [
            [
                m.get("member_id", ""),
                m.get("discipline", ""),
                m.get("member_type", ""),
                m.get("member_code", ""),
                m.get("level", ""),
                m.get("basis_status", ""),
                m.get("execution_status", ""),
                m.get("grid_ref", ""),
                m.get("clear_span", 0.0),
                m.get("section_width", 0.0),
                m.get("section_depth", 0.0),
                m.get("slab_type", ""),
                m.get("thickness", 0.0),
                m.get("type_ref", ""),
                m.get("H_to_top_of_beam", 0.0),
                m.get("source_ref", ""),
                m.get("notes", ""),
            ]
            for m in project["takeoff"]["members"]
        ],
    )

    raw_segments = wb.create_sheet("Raw_Segments")
    raw_segment_headers = [
        "segment_id",
        "member_id",
        "segment_name",
        "length",
        "width",
        "depth",
        "height",
        "area",
        "volume",
        "effective_length",
        "effective_width",
        "effective_depth",
        "effective_volume",
        "override_note",
        "basis_status",
        "source_ref",
        "formula_text",
        "origin_x",
        "origin_y",
        "origin_z",
    ]
    for idx, header in enumerate(raw_segment_headers, start=1):
        raw_segments.cell(row=1, column=idx, value=header)
    _style_header(raw_segments, 1, len(raw_segment_headers))
    for r_idx, segment in enumerate(project["takeoff"]["segments"], start=2):
        raw_segments.cell(row=r_idx, column=1, value=segment.get("segment_id", ""))
        raw_segments.cell(row=r_idx, column=2, value=segment.get("member_id", ""))
        raw_segments.cell(row=r_idx, column=3, value=segment.get("segment_name", ""))
        raw_segments.cell(row=r_idx, column=4, value=segment.get("length", 0.0))
        raw_segments.cell(row=r_idx, column=5, value=segment.get("width", 0.0))
        raw_segments.cell(row=r_idx, column=6, value=segment.get("depth", 0.0))
        raw_segments.cell(row=r_idx, column=7, value=segment.get("height", 0.0))
        raw_segments.cell(row=r_idx, column=8, value=segment.get("area", 0.0))
        raw_segments.cell(row=r_idx, column=9, value=f"=D{r_idx}*E{r_idx}*F{r_idx}")
        raw_segments.cell(row=r_idx, column=10, value=segment.get("overrides", {}).get("length", segment.get("length", 0.0)))
        raw_segments.cell(row=r_idx, column=11, value=segment.get("overrides", {}).get("width", segment.get("width", 0.0)))
        raw_segments.cell(row=r_idx, column=12, value=segment.get("overrides", {}).get("depth", segment.get("depth", 0.0)))
        raw_segments.cell(row=r_idx, column=13, value=f"=J{r_idx}*K{r_idx}*L{r_idx}")
        raw_segments.cell(row=r_idx, column=14, value="; ".join(
            f"{note.get('field')}={note.get('value')} ({note.get('justification', '')})"
            for note in segment.get("override_notes", [])
        ))
        raw_segments.cell(row=r_idx, column=15, value=segment.get("basis_status", ""))
        raw_segments.cell(row=r_idx, column=16, value=segment.get("source_ref", ""))
        raw_segments.cell(row=r_idx, column=17, value=segment.get("formula_text", ""))
        raw_segments.cell(row=r_idx, column=18, value=segment.get("origin_x", 0.0))
        raw_segments.cell(row=r_idx, column=19, value=segment.get("origin_y", 0.0))
        raw_segments.cell(row=r_idx, column=20, value=segment.get("origin_z", 0.0))
    raw_segments.freeze_panes = "A2"
    _auto_width(raw_segments)

    raw_components = wb.create_sheet("Raw_Components")
    _write_table(
        raw_components,
        ["component_id", "member_id", "source_segment_id", "component_type", "spec", "qty", "unit", "loss_pct", "line_type", "rate_context", "basis_status", "source_ref", "abt_charged_override", "formula_text"],
        [
            [
                c.get("component_id", ""),
                c.get("member_id", ""),
                c.get("source_segment_id", ""),
                c.get("component_type", ""),
                c.get("spec", ""),
                c.get("qty", 0.0),
                c.get("unit", ""),
                c.get("loss_pct", 0.0),
                c.get("line_type", ""),
                c.get("rate_context", ""),
                c.get("basis_status", ""),
                c.get("source_ref", ""),
                c.get("abt_charged_override"),
                c.get("formula_text", ""),
            ]
            for c in project["takeoff"]["components"]
        ],
    )

    raw_candidates = wb.create_sheet("Raw_Candidates")
    _write_table(
        raw_candidates,
        [
            "candidate_id",
            "candidate_type",
            "candidate_source",
            "confirmation_status",
            "confirmed_component_id",
            "member_id",
            "component_type",
            "spec",
            "qty",
            "unit",
            "rate_context",
            "basis_status",
            "source_ref",
            "review_note",
        ],
        [
            [
                c.get("candidate_id", ""),
                c.get("candidate_type", ""),
                c.get("candidate_source", ""),
                c.get("confirmation_status", ""),
                c.get("confirmed_component_id", ""),
                c.get("proposed_component", {}).get("member_id", ""),
                c.get("proposed_component", {}).get("component_type", ""),
                c.get("proposed_component", {}).get("spec", ""),
                c.get("proposed_component", {}).get("qty", 0.0),
                c.get("proposed_component", {}).get("unit", ""),
                c.get("proposed_component", {}).get("rate_context", ""),
                c.get("proposed_component", {}).get("basis_status", ""),
                c.get("proposed_component", {}).get("source_ref", ""),
                c.get("review_note", ""),
            ]
            for c in project.get("candidates", {}).get("components", [])
        ],
    )

    calc_rows = wb.create_sheet("Calc_Rows")
    calc_headers = [
        "calc_row_id", "member_code", "desc", "line_type", "rate_context", "qty", "loss_pct", "qty_with_loss",
        "unit", "mat_rate", "lab_rate", "mat_total", "lab_total", "line_total", "basis_status", "source_ref", "formula_text",
    ]
    for idx, header in enumerate(calc_headers, start=1):
        calc_rows.cell(row=1, column=idx, value=header)
    _style_header(calc_rows, 1, len(calc_headers))
    for r_idx, row in enumerate(project.get("calc_graph", {}).get("rows", []), start=2):
        source_segment_id = row.get("source_segment_id", "")
        calc_rows.cell(row=r_idx, column=1, value=row.get("calc_row_id", ""))
        calc_rows.cell(row=r_idx, column=2, value=row.get("member_code", ""))
        calc_rows.cell(row=r_idx, column=3, value=row.get("desc", ""))
        calc_rows.cell(row=r_idx, column=4, value=row.get("line_type", ""))
        calc_rows.cell(row=r_idx, column=5, value=row.get("rate_context", ""))
        if source_segment_id and source_segment_id in seg_row_map:
            calc_rows.cell(row=r_idx, column=6, value=f"='Raw_Segments'!M{seg_row_map[source_segment_id]}")
        else:
            calc_rows.cell(row=r_idx, column=6, value=row.get("qty", 0.0))
        calc_rows.cell(row=r_idx, column=7, value=row.get("loss_pct", 0.0))
        calc_rows.cell(row=r_idx, column=8, value=f"=F{r_idx}*(1+G{r_idx})")
        calc_rows.cell(row=r_idx, column=9, value=row.get("unit", ""))
        calc_rows.cell(row=r_idx, column=10, value=row.get("material_rate", 0.0))
        calc_rows.cell(row=r_idx, column=11, value=row.get("labor_rate", 0.0))
        calc_rows.cell(row=r_idx, column=12, value=f"=H{r_idx}*J{r_idx}")
        calc_rows.cell(row=r_idx, column=13, value=f"=H{r_idx}*K{r_idx}")
        if row.get("line_type") == "ABT" and row.get("abt_charged_override") not in (None, ""):
            calc_rows.cell(row=r_idx, column=14, value=float(row.get("abt_charged_override")))
        else:
            calc_rows.cell(row=r_idx, column=14, value=f"=L{r_idx}+M{r_idx}")
        calc_rows.cell(row=r_idx, column=15, value=row.get("basis_status", ""))
        calc_rows.cell(row=r_idx, column=16, value=row.get("source_ref", ""))
        calc_rows.cell(row=r_idx, column=17, value=row.get("formula_text", ""))
    calc_rows.freeze_panes = "A2"
    _auto_width(calc_rows)

    link_sheet = wb.create_sheet("Link_Sheet")
    _write_table(
        link_sheet,
        ["calc_row_id", "detail_row", "detail_line_total_ref", "raw_seg_row", "source_ref", "basis_status"],
        [
            [
                row.get("calc_row_id", ""),
                idx + 1,
                f"=Detail_BOQ!N{idx + 2}",
                seg_row_map.get(row.get("source_segment_id", ""), ""),
                row.get("source_ref", ""),
                row.get("basis_status", ""),
            ]
            for idx, row in enumerate(project.get("calc_graph", {}).get("rows", []))
        ],
    )

    detail = wb.create_sheet("Detail_BOQ")
    detail_headers = [
        "boq_line_id", "category", "description", "line_type", "basis_status", "qty", "unit",
        "mat_rate", "lab_rate", "mat_total", "lab_total", "machinery_total", "source_ref", "line_total", "calc_row_ref",
    ]
    for idx, header in enumerate(detail_headers, start=1):
        detail.cell(row=1, column=idx, value=header)
    _style_header(detail, 1, len(detail_headers))
    for r_idx, row in enumerate(project.get("calc_graph", {}).get("boq_lines", []), start=2):
        calc_ref = row.get("calc_row_ref", "")
        match_row = next((i for i, item in enumerate(project.get("calc_graph", {}).get("rows", []), start=2) if item.get("calc_row_id") == calc_ref), None)
        detail.cell(row=r_idx, column=1, value=row.get("boq_line_id", ""))
        detail.cell(row=r_idx, column=2, value=row.get("category", ""))
        detail.cell(row=r_idx, column=3, value=row.get("description", ""))
        detail.cell(row=r_idx, column=4, value=row.get("line_type", ""))
        detail.cell(row=r_idx, column=5, value=row.get("basis_status", ""))
        if match_row:
            detail.cell(row=r_idx, column=6, value=f"=Calc_Rows!H{match_row}")
            detail.cell(row=r_idx, column=7, value=f"=Calc_Rows!I{match_row}")
            detail.cell(row=r_idx, column=8, value=f"=Calc_Rows!J{match_row}")
            detail.cell(row=r_idx, column=9, value=f"=Calc_Rows!K{match_row}")
            detail.cell(row=r_idx, column=10, value=f"=Calc_Rows!L{match_row}")
            detail.cell(row=r_idx, column=11, value=f"=Calc_Rows!M{match_row}")
            detail.cell(row=r_idx, column=12, value=0)
            detail.cell(row=r_idx, column=13, value=f"=Calc_Rows!P{match_row}")
            detail.cell(row=r_idx, column=14, value=f"=Calc_Rows!N{match_row}")
            detail.cell(row=r_idx, column=15, value=calc_ref)
        else:
            detail.cell(row=r_idx, column=6, value=row.get("qty", 0.0))
            detail.cell(row=r_idx, column=7, value=row.get("unit", ""))
            detail.cell(row=r_idx, column=8, value=row.get("mat_rate", 0.0))
            detail.cell(row=r_idx, column=9, value=row.get("lab_rate", 0.0))
            detail.cell(row=r_idx, column=10, value=f"=F{r_idx}*H{r_idx}")
            detail.cell(row=r_idx, column=11, value=f"=F{r_idx}*I{r_idx}")
            detail.cell(row=r_idx, column=12, value=0)
            detail.cell(row=r_idx, column=13, value=row.get("source_ref", ""))
            detail.cell(row=r_idx, column=14, value=f"=J{r_idx}+K{r_idx}+L{r_idx}")
            detail.cell(row=r_idx, column=15, value=calc_ref)
    detail.freeze_panes = "A2"
    _auto_width(detail)

    summary_sheet = wb.create_sheet("Summary")
    summary_sheet["A1"] = "Internal Trace Summary"
    summary_sheet["A1"].font = FONT_BOLD
    summary_sheet["A3"] = "Project"
    summary_sheet["B3"] = project["project"].get("name", "")
    summary_sheet["A4"] = "Client"
    summary_sheet["B4"] = project["project"].get("client", "")
    summary_sheet["A5"] = "Direct Cost (workbook trace)"
    summary_sheet["B5"] = "=B9"
    summary_sheet["A6"] = "Factor F (kernel computed - update when policy changes)"
    summary_sheet["B6"] = summary.get("F", 0.0)
    summary_sheet["A7"] = "Final Bid (workbook trace)"
    summary_sheet["B7"] = "=B5*B6"
    summary_sheet["A9"] = "Detail sum by workbook"
    summary_sheet["B9"] = f"=SUM(Detail_BOQ!N2:N{max(2, len(project.get('calc_graph', {}).get('boq_lines', [])) + 1)})"
    summary_sheet["A11"] = "Category"
    summary_sheet["B11"] = "Line Total"
    _style_header(summary_sheet, 11, 2)
    categories = sorted({row.get("category", "Other") for row in project.get("calc_graph", {}).get("boq_lines", [])})
    start_row = 12
    for idx, category in enumerate(categories, start=start_row):
        summary_sheet.cell(row=idx, column=1, value=category)
        summary_sheet.cell(row=idx, column=2, value=f'=SUMIF(Detail_BOQ!B:B,A{idx},Detail_BOQ!N:N)')
    _auto_width(summary_sheet)

    review_sheet = wb.create_sheet("Review")
    _write_table(
        review_sheet,
        ["flag_id", "severity", "flag_type", "target_ref", "message", "export_rule", "resolution_status", "resolution_kind", "ack_comment", "ack_timestamp"],
        [
            [
                flag.get("flag_id", ""),
                flag.get("severity", ""),
                flag.get("flag_type", ""),
                flag.get("target_ref", ""),
                flag.get("message", ""),
                flag.get("export_rule", ""),
                flag.get("resolution_status", ""),
                flag.get("resolution_kind", ""),
                flag.get("ack_comment", ""),
                flag.get("ack_timestamp", ""),
            ]
            for flag in review_flags
        ],
    )

    for ws in wb.worksheets:
        if ws.title in {"Summary"}:
            continue
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and cell.row != 1:
                    cell.font = FONT

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return str(output)
