from __future__ import annotations

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from .acceptance_checker import evaluate_project_acceptance


def add_acceptance_sheet(project_id: str, workbook_path: str) -> None:
    evaluation = evaluate_project_acceptance(project_id)
    wb = openpyxl.load_workbook(workbook_path)
    
    if "Acceptance" in wb.sheetnames:
        del wb["Acceptance"]
        
    ws = wb.create_sheet("Acceptance", 0) # Make it the first sheet
    
    # Styles
    header_font = Font(name="Calibri", size=14, bold=True)
    normal_font = Font(name="Calibri", size=11)
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    
    ws["A1"] = "Acceptance Evaluation Report"
    ws["A1"].font = header_font
    
    ws["A3"] = "Overall Status"
    ws["B3"] = "PASSED" if evaluation["ok"] else "FAILED"
    ws["B3"].fill = pass_fill if evaluation["ok"] else fail_fill
    ws["B3"].font = Font(bold=True)
    
    row = 5
    ws.cell(row=row, column=1, value="Criterion").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Result").font = Font(bold=True)
    row += 1
    
    for key, val in evaluation["criteria"].items():
        ws.cell(row=row, column=1, value=key).font = normal_font
        res_cell = ws.cell(row=row, column=2, value="OK" if val else "FAIL")
        res_cell.fill = pass_fill if val else fail_fill
        res_cell.alignment = Alignment(horizontal="center")
        row += 1
        
    row += 1
    if evaluation.get("override"):
        ws.cell(row=row, column=1, value="Acceptance Override").font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="Author")
        ws.cell(row=row, column=2, value=evaluation["override"].get("author"))
        row += 1
        ws.cell(row=row, column=1, value="Justification")
        ws.cell(row=row, column=2, value=evaluation["override"].get("justification"))
        row += 1
        ws.cell(row=row, column=1, value="Timestamp")
        ws.cell(row=row, column=2, value=evaluation["override"].get("timestamp"))
        
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50
    
    wb.save(workbook_path)
